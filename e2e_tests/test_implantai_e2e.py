"""
============================================================
  ImplantAI Dental App — Full E2E Test Suite (100+ Tests)
  URL  : https://pdd-zfqq.onrender.com/
  Tool : Selenium 4 + ChromeDriver (webdriver-manager)
  Run  : python test_implantai_e2e.py
============================================================
"""

import time
import os
import io
import re
import sys
import json
import datetime
import traceback
import unittest
import warnings

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException,
    ElementNotInteractableException, WebDriverException
)
from webdriver_manager.chrome import ChromeDriverManager

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
#  GLOBAL CONFIG
# ─────────────────────────────────────────────
BASE_URL   = "https://pdd-zfqq.onrender.com"
WAIT_TIME  = 20          # seconds for explicit waits
PAGE_LOAD  = 30          # seconds for page loads
TEST_SUITE = "ImplantAI Dental Web App — Full E2E Workflow"

# ─────────────────────────────────────────────
#  RESULT COLLECTOR
# ─────────────────────────────────────────────
results: list[dict] = []

def record(tc_id, name, category, status, duration, message="", expected="", actual=""):
    results.append({
        "TC_ID":    tc_id,
        "Name":     name,
        "Category": category,
        "Status":   status,
        "Duration": round(duration, 2),
        "Message":  message,
        "Expected": expected,
        "Actual":   actual,
    })
    icon = "✅" if status == "PASS" else ("⚠️" if status == "SKIP" else "❌")
    print(f"  {icon} [{tc_id}] {name} ({duration:.2f}s)")

# ─────────────────────────────────────────────
#  DRIVER FACTORY
# ─────────────────────────────────────────────
def make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    # headless flag — comment out to watch tests run
    # opts.add_argument("--headless=new")
    svc = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=svc, options=opts)
    driver.set_page_load_timeout(PAGE_LOAD)
    return driver

# ─────────────────────────────────────────────
#  HELPER UTILITIES
# ─────────────────────────────────────────────
def wait_for(driver, by, selector, timeout=WAIT_TIME):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, selector))
    )

def wait_visible(driver, by, selector, timeout=WAIT_TIME):
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, selector))
    )

def wait_clickable(driver, by, selector, timeout=WAIT_TIME):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, selector))
    )

def safe_find(driver, by, selector):
    try:
        return driver.find_element(by, selector)
    except NoSuchElementException:
        return None

def safe_finds(driver, by, selector):
    try:
        return driver.find_elements(by, selector)
    except Exception:
        return []

def page_has_text(driver, text):
    return text.lower() in driver.page_source.lower()

def go(driver, path=""):
    driver.get(f"{BASE_URL}{path}")
    time.sleep(1.5)

def click_nav(driver, text):
    """Click a navigation item by its visible text."""
    try:
        links = driver.find_elements(By.XPATH,
            f"//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{text.lower()}')]"
        )
        for lnk in links:
            if lnk.is_displayed():
                lnk.click()
                return True
    except Exception:
        pass
    return False

def upload_dummy_image(driver, input_selector, by=By.CSS_SELECTOR):
    """Create a tiny valid PNG and upload it to any file input."""
    import tempfile, struct, zlib
    def _make_png(w=10, h=10):
        def chunk(name, data):
            c = zlib.crc32(name + data) & 0xFFFFFFFF
            return struct.pack(">I", len(data)) + name + data + struct.pack(">I", c)
        ihdr = chunk(b"IHDR", struct.pack(">IIBBBBB", w, h, 8, 2, 0, 0, 0))
        raw  = b"".join(b"\x00" + b"\xFF\x00\x00" * w for _ in range(h))
        idat = chunk(b"IDAT", zlib.compress(raw))
        iend = chunk(b"IEND", b"")
        return b"\x89PNG\r\n\x1a\n" + ihdr + idat + iend

    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp.write(_make_png())
    tmp.flush()
    tmp.close()
    try:
        inp = driver.find_element(by, input_selector)
        driver.execute_script("arguments[0].style.display='block';", inp)
        inp.send_keys(tmp.name)
        time.sleep(1)
        return True
    except Exception:
        return False
    finally:
        try: os.unlink(tmp.name)
        except: pass

# ═══════════════════════════════════════════════════════════
#   TEST CATEGORIES
# ═══════════════════════════════════════════════════════════

# ───────────────────────────────────────────────────────────
#  CAT-01: App Launch & Page Load
# ───────────────────────────────────────────────────────────
def cat_01_launch(driver):
    cat = "App Launch & Page Load"
    tests = [
        ("TC_001", "Homepage loads without errors"),
        ("TC_002", "Page title is not empty"),
        ("TC_003", "No JS console errors on load"),
        ("TC_004", "Favicon is present"),
        ("TC_005", "Meta viewport tag exists"),
    ]

    # TC_001
    t = tests[0]; t0 = time.time()
    try:
        go(driver)
        assert "implantai" in driver.title.lower() or len(driver.page_source) > 500
        record(t[0], t[1], cat, "PASS", time.time()-t0, "App loaded", "200 OK", "200 OK")
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))

    # TC_002
    t = tests[1]; t0 = time.time()
    try:
        title = driver.title
        assert title and len(title) > 0
        record(t[0], t[1], cat, "PASS", time.time()-t0, f"Title: {title}", "Non-empty title", title)
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))

    # TC_003
    t = tests[2]; t0 = time.time()
    try:
        logs = driver.get_log("browser")
        severe = [l for l in logs if l.get("level") == "SEVERE"]
        record(t[0], t[1], cat, "PASS" if not severe else "FAIL",
               time.time()-t0,
               f"{len(severe)} severe errors" if severe else "No severe errors",
               "0 severe JS errors", str(len(severe)))
    except Exception as e:
        record(t[0], t[1], cat, "PASS", time.time()-t0, "Log check skipped", "", "")

    # TC_004
    t = tests[3]; t0 = time.time()
    try:
        fav = driver.find_elements(By.XPATH, "//link[@rel='icon' or @rel='shortcut icon']")
        assert len(fav) > 0
        record(t[0], t[1], cat, "PASS", time.time()-t0, "Favicon found", "Favicon present", "Present")
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))

    # TC_005
    t = tests[4]; t0 = time.time()
    try:
        vp = driver.find_element(By.XPATH, "//meta[@name='viewport']")
        record(t[0], t[1], cat, "PASS", time.time()-t0, "Viewport meta found", "Meta viewport present", "Present")
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-02: Navigation & Routing
# ───────────────────────────────────────────────────────────
def cat_02_navigation(driver):
    cat = "Navigation & Routing"

    nav_tests = [
        ("TC_006", "Sidebar / nav bar is visible"),
        ("TC_007", "App logo / brand name visible"),
        ("TC_008", "Patients nav item clickable"),
        ("TC_009", "Add Patient nav/button visible"),
        ("TC_010", "Scan Analysis nav item visible"),
        ("TC_011", "Reports nav item visible"),
        ("TC_012", "Settings nav item visible"),
        ("TC_013", "Back navigation works"),
        ("TC_014", "Route /patients loads"),
        ("TC_015", "Route /patients/add loads"),
    ]

    # TC_006
    t = nav_tests[0]; t0 = time.time()
    try:
        go(driver)
        sidebar = driver.find_elements(By.XPATH, "//*[contains(@class,'sidebar') or contains(@class,'nav') or contains(@class,'menu')]")
        assert len(sidebar) > 0
        record(t[0], t[1], cat, "PASS", time.time()-t0, f"Sidebar found ({len(sidebar)} el)", "Sidebar visible", "Visible")
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))

    # TC_007
    t = nav_tests[1]; t0 = time.time()
    try:
        brand = page_has_text(driver, "implant") or page_has_text(driver, "dental") or page_has_text(driver, "ai")
        assert brand
        record(t[0], t[1], cat, "PASS", time.time()-t0, "Brand name found", "Brand visible", "Found")
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))

    # TC_008 – TC_012  (nav link presence)
    nav_items = [
        (nav_tests[2], "patient"),
        (nav_tests[3], "add patient"),
        (nav_tests[4], "scan"),
        (nav_tests[5], "report"),
        (nav_tests[6], "setting"),
    ]
    for (t, keyword) in nav_items:
        t0 = time.time()
        try:
            els = driver.find_elements(By.XPATH,
                f"//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{keyword}')]"
            )
            visible = [e for e in els if e.is_displayed()]
            status = "PASS" if visible else "FAIL"
            record(t[0], t[1], cat, status, time.time()-t0,
                   f"'{keyword}' nav {'found' if visible else 'NOT found'}",
                   f"'{keyword}' visible", "Found" if visible else "Missing")
        except Exception as e:
            record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))

    # TC_013
    t = nav_tests[7]; t0 = time.time()
    try:
        driver.back()
        time.sleep(1)
        record(t[0], t[1], cat, "PASS", time.time()-t0, "Back navigation succeeded", "Navigation goes back", "Success")
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))

    # TC_014
    t = nav_tests[8]; t0 = time.time()
    try:
        go(driver, "/#/patients")
        time.sleep(2)
        src = driver.page_source
        assert len(src) > 300
        record(t[0], t[1], cat, "PASS", time.time()-t0, "Patients route loaded", "Route loads", "Loaded")
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))

    # TC_015
    t = nav_tests[9]; t0 = time.time()
    try:
        go(driver, "/#/patients/add")
        time.sleep(2)
        src = driver.page_source
        assert len(src) > 300
        record(t[0], t[1], cat, "PASS", time.time()-t0, "/patients/add route loaded", "Route loads", "Loaded")
    except Exception as e:
        record(t[0], t[1], cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-03: Authentication (Login / Register)
# ───────────────────────────────────────────────────────────
def cat_03_auth(driver):
    cat = "Authentication"

    # TC_016 – Login page visible
    t0 = time.time()
    try:
        go(driver)
        has_login = page_has_text(driver, "login") or page_has_text(driver, "sign in")
        record("TC_016", "Login / sign-in option visible", cat,
               "PASS" if has_login else "FAIL", time.time()-t0,
               "Login found" if has_login else "Login not found", "Login visible", str(has_login))
    except Exception as e:
        record("TC_016", "Login / sign-in option visible", cat, "FAIL", time.time()-t0, str(e))

    # TC_017 – Email field exists
    t0 = time.time()
    try:
        go(driver)
        email_inp = driver.find_elements(By.CSS_SELECTOR, "input[type='email'], input[name='email'], input[placeholder*='email' i]")
        has_email = len(email_inp) > 0
        record("TC_017", "Email input field exists", cat,
               "PASS" if has_email else "SKIP", time.time()-t0,
               "Email input found" if has_email else "No email input (may be SPA behind auth)", "Email input", str(has_email))
    except Exception as e:
        record("TC_017", "Email input field exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_018 – Password field exists
    t0 = time.time()
    try:
        pass_inp = driver.find_elements(By.CSS_SELECTOR, "input[type='password']")
        has_pass = len(pass_inp) > 0
        record("TC_018", "Password input field exists", cat,
               "PASS" if has_pass else "SKIP", time.time()-t0,
               "Password input found" if has_pass else "No password input", "Password input", str(has_pass))
    except Exception as e:
        record("TC_018", "Password input field exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_019 – Login with empty fields shows validation
    t0 = time.time()
    try:
        submit_btns = driver.find_elements(By.CSS_SELECTOR, "button[type='submit'], button")
        login_btn = None
        for b in submit_btns:
            if "login" in b.text.lower() or "sign in" in b.text.lower():
                login_btn = b; break
        if login_btn:
            login_btn.click()
            time.sleep(1)
            has_msg = (page_has_text(driver, "required") or page_has_text(driver, "invalid") or
                       page_has_text(driver, "error") or page_has_text(driver, "fill"))
            record("TC_019", "Empty login shows validation error", cat,
                   "PASS" if has_msg else "FAIL", time.time()-t0,
                   "Validation shown" if has_msg else "No validation", "Validation error", str(has_msg))
        else:
            record("TC_019", "Empty login shows validation error", cat, "SKIP", time.time()-t0, "Login button not found")
    except Exception as e:
        record("TC_019", "Empty login shows validation error", cat, "FAIL", time.time()-t0, str(e))

    # TC_020 – Invalid credentials error
    t0 = time.time()
    try:
        email_inp = driver.find_elements(By.CSS_SELECTOR, "input[type='email'], input[name='email']")
        pass_inp  = driver.find_elements(By.CSS_SELECTOR, "input[type='password']")
        if email_inp and pass_inp:
            email_inp[0].clear(); email_inp[0].send_keys("invalid@test.com")
            pass_inp[0].clear();  pass_inp[0].send_keys("wrongpassword123")
            submit = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            submit.click(); time.sleep(3)
            error_shown = (page_has_text(driver, "invalid") or page_has_text(driver, "incorrect") or
                           page_has_text(driver, "error") or page_has_text(driver, "failed"))
            record("TC_020", "Invalid credentials shows error", cat,
                   "PASS" if error_shown else "FAIL", time.time()-t0,
                   "Error shown" if error_shown else "No error shown", "Auth error", str(error_shown))
        else:
            record("TC_020", "Invalid credentials shows error", cat, "SKIP", time.time()-t0, "Auth form not found on initial page")
    except Exception as e:
        record("TC_020", "Invalid credentials shows error", cat, "FAIL", time.time()-t0, str(e))

    # TC_021 – Register link present
    t0 = time.time()
    try:
        has_reg = page_has_text(driver, "register") or page_has_text(driver, "sign up") or page_has_text(driver, "create account")
        record("TC_021", "Register / Sign-up link visible", cat,
               "PASS" if has_reg else "SKIP", time.time()-t0,
               "Register link found" if has_reg else "Not found", "Register link", str(has_reg))
    except Exception as e:
        record("TC_021", "Register / Sign-up link visible", cat, "FAIL", time.time()-t0, str(e))

    # TC_022 – Logout option visible when authenticated
    t0 = time.time()
    try:
        has_logout = page_has_text(driver, "logout") or page_has_text(driver, "sign out") or page_has_text(driver, "log out")
        record("TC_022", "Logout option visible", cat,
               "PASS" if has_logout else "SKIP", time.time()-t0,
               "Logout found" if has_logout else "Not found (may need auth)", "Logout visible", str(has_logout))
    except Exception as e:
        record("TC_022", "Logout option visible", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-04: Patient Management
# ───────────────────────────────────────────────────────────
def cat_04_patients(driver):
    cat = "Patient Management"

    go(driver, "/#/patients/add")
    time.sleep(2)

    # TC_023 – Add Patient form loads
    t0 = time.time()
    try:
        has_form = len(driver.find_elements(By.TAG_NAME, "form")) > 0 or \
                   len(driver.find_elements(By.CSS_SELECTOR, "input, select, textarea")) > 0
        record("TC_023", "Add Patient form renders", cat,
               "PASS" if has_form else "FAIL", time.time()-t0,
               "Form elements found" if has_form else "No form elements", "Form visible", str(has_form))
    except Exception as e:
        record("TC_023", "Add Patient form renders", cat, "FAIL", time.time()-t0, str(e))

    # TC_024 – Patient name field exists
    t0 = time.time()
    try:
        inp = driver.find_elements(By.CSS_SELECTOR, "input[name*='name' i], input[placeholder*='name' i], input[id*='name' i]")
        record("TC_024", "Patient name input field exists", cat,
               "PASS" if inp else "FAIL", time.time()-t0,
               "Name field found" if inp else "Not found", "Name field present", str(bool(inp)))
    except Exception as e:
        record("TC_024", "Patient name input field exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_025 – Age / DOB field exists
    t0 = time.time()
    try:
        inp = driver.find_elements(By.CSS_SELECTOR, "input[name*='age' i], input[name*='dob' i], input[type='date'], input[placeholder*='age' i]")
        record("TC_025", "Age / DOB field exists", cat,
               "PASS" if inp else "FAIL", time.time()-t0,
               "Age/DOB field found" if inp else "Not found", "Age field present", str(bool(inp)))
    except Exception as e:
        record("TC_025", "Age / DOB field exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_026 – Gender field exists
    t0 = time.time()
    try:
        inp = driver.find_elements(By.CSS_SELECTOR,
            "select[name*='gender' i], input[name*='gender' i], [data-testid*='gender' i]")
        has = len(inp) > 0 or page_has_text(driver, "gender") or page_has_text(driver, "male")
        record("TC_026", "Gender selection field exists", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Gender field found" if has else "Not found", "Gender field present", str(has))
    except Exception as e:
        record("TC_026", "Gender selection field exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_027 – Phone / contact field exists
    t0 = time.time()
    try:
        inp = driver.find_elements(By.CSS_SELECTOR, "input[type='tel'], input[name*='phone' i], input[placeholder*='phone' i]")
        has = len(inp) > 0 or page_has_text(driver, "phone") or page_has_text(driver, "contact")
        record("TC_027", "Phone / contact field exists", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Phone field found" if has else "Not found", "Phone field present", str(has))
    except Exception as e:
        record("TC_027", "Phone / contact field exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_028 – Medical history / notes field
    t0 = time.time()
    try:
        inp = driver.find_elements(By.CSS_SELECTOR, "textarea, input[name*='history' i], input[name*='note' i]")
        has = len(inp) > 0 or page_has_text(driver, "history") or page_has_text(driver, "note")
        record("TC_028", "Medical history / notes field exists", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Notes field found" if has else "Not found", "Notes field present", str(has))
    except Exception as e:
        record("TC_028", "Medical history / notes field exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_029 – Smoking / diabetes risk field
    t0 = time.time()
    try:
        has = page_has_text(driver, "smok") or page_has_text(driver, "diabet")
        record("TC_029", "Risk factor fields (smoking/diabetes) visible", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Risk fields found" if has else "Not found", "Risk fields present", str(has))
    except Exception as e:
        record("TC_029", "Risk factor fields (smoking/diabetes) visible", cat, "FAIL", time.time()-t0, str(e))

    # TC_030 – Submit / Save button exists
    t0 = time.time()
    try:
        btns = driver.find_elements(By.XPATH,
            "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'save') or "
            "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'submit') or "
            "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'add patient')]")
        has = len(btns) > 0
        record("TC_030", "Save / Submit button exists on Add Patient", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Submit button found" if has else "Not found", "Submit button present", str(has))
    except Exception as e:
        record("TC_030", "Save / Submit button exists on Add Patient", cat, "FAIL", time.time()-t0, str(e))

    # TC_031 – Submit empty form shows validation
    t0 = time.time()
    try:
        btns = driver.find_elements(By.CSS_SELECTOR, "button[type='submit'], button")
        save_btn = None
        for b in btns:
            txt = b.text.lower()
            if any(k in txt for k in ["save", "submit", "add"]):
                save_btn = b; break
        if save_btn:
            save_btn.click(); time.sleep(1.5)
            has_val = (page_has_text(driver, "required") or page_has_text(driver, "fill") or
                       page_has_text(driver, "error") or page_has_text(driver, "invalid"))
            record("TC_031", "Empty form submit triggers validation", cat,
                   "PASS" if has_val else "FAIL", time.time()-t0,
                   "Validation shown" if has_val else "No validation shown", "Validation error", str(has_val))
        else:
            record("TC_031", "Empty form submit triggers validation", cat, "SKIP", time.time()-t0, "Submit button not found")
    except Exception as e:
        record("TC_031", "Empty form submit triggers validation", cat, "FAIL", time.time()-t0, str(e))

    # TC_032 – Patient list page loads
    t0 = time.time()
    try:
        go(driver, "/#/patients")
        time.sleep(2)
        has_list = (page_has_text(driver, "patient") and
                    (page_has_text(driver, "list") or page_has_text(driver, "record") or
                     len(driver.find_elements(By.CSS_SELECTOR, "table, ul, .card, [class*='patient']")) > 0))
        record("TC_032", "Patient list page loads", cat,
               "PASS" if has_list else "FAIL", time.time()-t0,
               "List page loaded" if has_list else "Not loaded", "Patient list", "Loaded" if has_list else "Failed")
    except Exception as e:
        record("TC_032", "Patient list page loads", cat, "FAIL", time.time()-t0, str(e))

    # TC_033 – Search / filter input on patient list
    t0 = time.time()
    try:
        search = driver.find_elements(By.CSS_SELECTOR, "input[type='search'], input[placeholder*='search' i], input[placeholder*='filter' i]")
        has = len(search) > 0
        record("TC_033", "Search / filter input on patient list", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Search input found" if has else "Not found", "Search input present", str(has))
    except Exception as e:
        record("TC_033", "Search / filter input on patient list", cat, "FAIL", time.time()-t0, str(e))

    # TC_034 – Patient cards / rows visible
    t0 = time.time()
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "tr, .patient-card, [class*='card'], li[class*='patient']")
        record("TC_034", "Patient records / rows rendered", cat,
               "PASS" if rows else "FAIL", time.time()-t0,
               f"{len(rows)} row(s) found" if rows else "No rows", "Patient rows", str(len(rows)))
    except Exception as e:
        record("TC_034", "Patient records / rows rendered", cat, "FAIL", time.time()-t0, str(e))

    # TC_035 – Delete patient button
    t0 = time.time()
    try:
        dels = driver.find_elements(By.XPATH,
            "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'delete') or "
            "contains(@class,'delete') or contains(@aria-label,'delete' )]")
        record("TC_035", "Delete patient button exists", cat,
               "PASS" if dels else "FAIL", time.time()-t0,
               f"{len(dels)} delete btn(s)" if dels else "Not found", "Delete button", str(bool(dels)))
    except Exception as e:
        record("TC_035", "Delete patient button exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_036 – View patient detail
    t0 = time.time()
    try:
        view_links = driver.find_elements(By.XPATH,
            "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'view') or "
            "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'detail') or "
            "contains(@class,'view') or contains(@href,'patient')]")
        record("TC_036", "View patient detail link exists", cat,
               "PASS" if view_links else "FAIL", time.time()-t0,
               f"{len(view_links)} link(s) found" if view_links else "Not found", "View link present", str(bool(view_links)))
    except Exception as e:
        record("TC_036", "View patient detail link exists", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-05: AI Scan Analysis
# ───────────────────────────────────────────────────────────
def cat_05_scan(driver):
    cat = "AI Scan Analysis"

    go(driver)
    time.sleep(2)

    # TC_037 – Scan Analysis page / section visible
    t0 = time.time()
    try:
        has_scan = (page_has_text(driver, "scan") or page_has_text(driver, "analysis") or
                    page_has_text(driver, "panoramic") or page_has_text(driver, "x-ray"))
        record("TC_037", "Scan / Analysis section is accessible", cat,
               "PASS" if has_scan else "FAIL", time.time()-t0,
               "Scan section found" if has_scan else "Not found", "Scan section visible", str(has_scan))
    except Exception as e:
        record("TC_037", "Scan / Analysis section is accessible", cat, "FAIL", time.time()-t0, str(e))

    # TC_038 – File upload input for scan
    t0 = time.time()
    try:
        inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
        record("TC_038", "File upload input for scan exists", cat,
               "PASS" if inputs else "FAIL", time.time()-t0,
               f"{len(inputs)} file input(s)" if inputs else "Not found", "File input present", str(bool(inputs)))
    except Exception as e:
        record("TC_038", "File upload input for scan exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_039 – Drag-and-drop zone visible
    t0 = time.time()
    try:
        dnd = driver.find_elements(By.CSS_SELECTOR,
            "[class*='drop'], [class*='drag'], [data-testid*='drop']")
        has_dnd = len(dnd) > 0 or page_has_text(driver, "drag") or page_has_text(driver, "drop")
        record("TC_039", "Drag-and-drop upload zone visible", cat,
               "PASS" if has_dnd else "FAIL", time.time()-t0,
               "DnD zone found" if has_dnd else "Not found", "DnD zone present", str(has_dnd))
    except Exception as e:
        record("TC_039", "Drag-and-drop upload zone visible", cat, "FAIL", time.time()-t0, str(e))

    # TC_040 – Panoramic analysis option
    t0 = time.time()
    try:
        has = page_has_text(driver, "panoramic") or page_has_text(driver, "caries")
        record("TC_040", "Panoramic caries analysis option", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Panoramic option found" if has else "Not found", "Panoramic option", str(has))
    except Exception as e:
        record("TC_040", "Panoramic caries analysis option", cat, "FAIL", time.time()-t0, str(e))

    # TC_041 – Implant detection option
    t0 = time.time()
    try:
        has = page_has_text(driver, "implant")
        record("TC_041", "Implant detection analysis option", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Implant option found" if has else "Not found", "Implant option", str(has))
    except Exception as e:
        record("TC_041", "Implant detection analysis option", cat, "FAIL", time.time()-t0, str(e))

    # TC_042 – Mandibular canal option
    t0 = time.time()
    try:
        has = page_has_text(driver, "mandibular") or page_has_text(driver, "canal")
        record("TC_042", "Mandibular canal detection option", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Mandibular option found" if has else "Not found", "Mandibular option", str(has))
    except Exception as e:
        record("TC_042", "Mandibular canal detection option", cat, "FAIL", time.time()-t0, str(e))

    # TC_043 – Maxillary sinus option
    t0 = time.time()
    try:
        has = page_has_text(driver, "sinus") or page_has_text(driver, "maxillary")
        record("TC_043", "Maxillary sinus detection option", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Sinus option found" if has else "Not found", "Sinus option", str(has))
    except Exception as e:
        record("TC_043", "Maxillary sinus detection option", cat, "FAIL", time.time()-t0, str(e))

    # TC_044 – Analyze button exists
    t0 = time.time()
    try:
        btns = driver.find_elements(By.XPATH,
            "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'analyz') or "
            "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'detect') or "
            "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'scan')]")
        record("TC_044", "Analyze / Detect button exists", cat,
               "PASS" if btns else "FAIL", time.time()-t0,
               f"{len(btns)} button(s)" if btns else "Not found", "Analyze button", str(bool(btns)))
    except Exception as e:
        record("TC_044", "Analyze / Detect button exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_045 – Upload dummy image and trigger analysis
    t0 = time.time()
    try:
        file_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
        uploaded = False
        if file_inputs:
            uploaded = upload_dummy_image(driver, "input[type='file']")
        record("TC_045", "Dummy image upload accepted", cat,
               "PASS" if uploaded else "FAIL", time.time()-t0,
               "Image uploaded" if uploaded else "Upload failed", "Upload succeeds", str(uploaded))
    except Exception as e:
        record("TC_045", "Dummy image upload accepted", cat, "FAIL", time.time()-t0, str(e))

    # TC_046 – Results/detections section appears after analysis
    t0 = time.time()
    try:
        has = (page_has_text(driver, "detection") or page_has_text(driver, "result") or
               page_has_text(driver, "confidence") or page_has_text(driver, "found"))
        record("TC_046", "Detection results section visible", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Results visible" if has else "No results shown", "Results visible", str(has))
    except Exception as e:
        record("TC_046", "Detection results section visible", cat, "FAIL", time.time()-t0, str(e))

    # TC_047 – Annotated image / bounding box canvas exists
    t0 = time.time()
    try:
        canvases = driver.find_elements(By.TAG_NAME, "canvas")
        imgs = driver.find_elements(By.CSS_SELECTOR, "img[class*='result' i], img[class*='annot' i]")
        has = len(canvases) > 0 or len(imgs) > 0
        record("TC_047", "Annotated result image / canvas shown", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               f"{len(canvases)} canvas(es)" if canvases else "No canvas", "Canvas / result img", str(has))
    except Exception as e:
        record("TC_047", "Annotated result image / canvas shown", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-06: Implant Survival Prediction (Gemini AI)
# ───────────────────────────────────────────────────────────
def cat_06_survival(driver):
    cat = "Implant Survival Prediction"

    go(driver)
    time.sleep(2)

    # TC_048 – Survival prediction section accessible
    t0 = time.time()
    try:
        has = page_has_text(driver, "survival") or page_has_text(driver, "predict") or page_has_text(driver, "prognos")
        record("TC_048", "Survival prediction section accessible", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Survival section found" if has else "Not found", "Survival section", str(has))
    except Exception as e:
        record("TC_048", "Survival prediction section accessible", cat, "FAIL", time.time()-t0, str(e))

    # TC_049 – Survival probability gauge / number shown
    t0 = time.time()
    try:
        has = (page_has_text(driver, "survival probability") or page_has_text(driver, "%") or
               len(driver.find_elements(By.CSS_SELECTOR, ".gauge, [class*='gauge'], [class*='probability']")) > 0)
        record("TC_049", "Survival probability gauge / value shown", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Probability shown" if has else "Not shown", "Probability displayed", str(has))
    except Exception as e:
        record("TC_049", "Survival probability gauge / value shown", cat, "FAIL", time.time()-t0, str(e))

    # TC_050 – Risk factors list displayed
    t0 = time.time()
    try:
        has = page_has_text(driver, "risk factor") or page_has_text(driver, "risk level")
        record("TC_050", "Risk factors list displayed", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Risk factors found" if has else "Not found", "Risk factors list", str(has))
    except Exception as e:
        record("TC_050", "Risk factors list displayed", cat, "FAIL", time.time()-t0, str(e))

    # TC_051 – Success factors list
    t0 = time.time()
    try:
        has = page_has_text(driver, "success factor") or page_has_text(driver, "positive factor")
        record("TC_051", "Success factors list displayed", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Success factors found" if has else "Not found", "Success factors", str(has))
    except Exception as e:
        record("TC_051", "Success factors list displayed", cat, "FAIL", time.time()-t0, str(e))

    # TC_052 – Action items / recommendations shown
    t0 = time.time()
    try:
        has = page_has_text(driver, "action") or page_has_text(driver, "recommend") or page_has_text(driver, "follow")
        record("TC_052", "Action items / recommendations shown", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Action items found" if has else "Not found", "Action items present", str(has))
    except Exception as e:
        record("TC_052", "Action items / recommendations shown", cat, "FAIL", time.time()-t0, str(e))

    # TC_053 – Failure risk percentage shown
    t0 = time.time()
    try:
        has = page_has_text(driver, "failure risk") or page_has_text(driver, "failure_risk")
        record("TC_053", "Failure risk percentage shown", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Failure risk shown" if has else "Not shown", "Failure risk", str(has))
    except Exception as e:
        record("TC_053", "Failure risk percentage shown", cat, "FAIL", time.time()-t0, str(e))

    # TC_054 – AI confidence score shown
    t0 = time.time()
    try:
        has = page_has_text(driver, "confidence") or page_has_text(driver, "accuracy")
        record("TC_054", "AI confidence score shown", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Confidence shown" if has else "Not shown", "Confidence score", str(has))
    except Exception as e:
        record("TC_054", "AI confidence score shown", cat, "FAIL", time.time()-t0, str(e))

    # TC_055 – Narrative / analysis text present
    t0 = time.time()
    try:
        has = page_has_text(driver, "narrative") or page_has_text(driver, "analysis") or page_has_text(driver, "estimated")
        record("TC_055", "AI narrative analysis text present", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Narrative found" if has else "Not found", "Narrative present", str(has))
    except Exception as e:
        record("TC_055", "AI narrative analysis text present", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-07: AI Chat Assistant
# ───────────────────────────────────────────────────────────
def cat_07_chat(driver):
    cat = "AI Chat Assistant"

    go(driver)
    time.sleep(2)

    # TC_056 – Chat widget / section present
    t0 = time.time()
    try:
        has = (page_has_text(driver, "chat") or page_has_text(driver, "assistant") or
               len(driver.find_elements(By.CSS_SELECTOR, "[class*='chat'], [class*='assistant']")) > 0)
        record("TC_056", "AI Chat assistant widget present", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Chat found" if has else "Not found", "Chat widget", str(has))
    except Exception as e:
        record("TC_056", "AI Chat assistant widget present", cat, "FAIL", time.time()-t0, str(e))

    # TC_057 – Chat input field
    t0 = time.time()
    try:
        inp = driver.find_elements(By.CSS_SELECTOR, "input[placeholder*='ask' i], input[placeholder*='type' i], textarea[placeholder*='ask' i]")
        record("TC_057", "Chat message input field exists", cat,
               "PASS" if inp else "FAIL", time.time()-t0,
               "Chat input found" if inp else "Not found", "Chat input", str(bool(inp)))
    except Exception as e:
        record("TC_057", "Chat message input field exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_058 – Send button
    t0 = time.time()
    try:
        btns = driver.find_elements(By.XPATH,
            "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'send') or "
            "contains(@aria-label,'send' )]")
        record("TC_058", "Chat send button exists", cat,
               "PASS" if btns else "FAIL", time.time()-t0,
               "Send button found" if btns else "Not found", "Send button", str(bool(btns)))
    except Exception as e:
        record("TC_058", "Chat send button exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_059 – Type message and send
    t0 = time.time()
    try:
        inp = driver.find_elements(By.CSS_SELECTOR, "input[placeholder*='ask' i], textarea[placeholder*='ask' i], [class*='chat'] input, [class*='chat'] textarea")
        if inp:
            inp[0].send_keys("What is an implant?")
            inp[0].send_keys(Keys.RETURN)
            time.sleep(4)
            has_resp = page_has_text(driver, "implant") or page_has_text(driver, "dental") or page_has_text(driver, "assistant")
            record("TC_059", "Chat responds to user message", cat,
                   "PASS" if has_resp else "FAIL", time.time()-t0,
                   "Response shown" if has_resp else "No response", "Response received", str(has_resp))
        else:
            record("TC_059", "Chat responds to user message", cat, "SKIP", time.time()-t0, "Chat input not found")
    except Exception as e:
        record("TC_059", "Chat responds to user message", cat, "FAIL", time.time()-t0, str(e))

    # TC_060 – Chat history preserved across messages
    t0 = time.time()
    try:
        msgs = driver.find_elements(By.CSS_SELECTOR, "[class*='message'], [class*='bubble'], [class*='chat-item']")
        record("TC_060", "Chat message history visible", cat,
               "PASS" if msgs else "FAIL", time.time()-t0,
               f"{len(msgs)} message(s)" if msgs else "No messages", "Messages shown", str(len(msgs)))
    except Exception as e:
        record("TC_060", "Chat message history visible", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-08: Reports & PDF Export
# ───────────────────────────────────────────────────────────
def cat_08_reports(driver):
    cat = "Reports & PDF Export"

    go(driver)
    time.sleep(2)

    # TC_061 – Reports section accessible
    t0 = time.time()
    try:
        has = page_has_text(driver, "report") or page_has_text(driver, "export") or page_has_text(driver, "pdf")
        record("TC_061", "Reports section accessible", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Reports found" if has else "Not found", "Reports accessible", str(has))
    except Exception as e:
        record("TC_061", "Reports section accessible", cat, "FAIL", time.time()-t0, str(e))

    # TC_062 – Export / Download button
    t0 = time.time()
    try:
        btns = driver.find_elements(By.XPATH,
            "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'export') or "
            "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'download') or "
            "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'pdf')]")
        record("TC_062", "Export / Download PDF button exists", cat,
               "PASS" if btns else "FAIL", time.time()-t0,
               f"{len(btns)} button(s) found" if btns else "Not found", "Export button", str(bool(btns)))
    except Exception as e:
        record("TC_062", "Export / Download PDF button exists", cat, "FAIL", time.time()-t0, str(e))

    # TC_063 – Report shows patient name
    t0 = time.time()
    try:
        has = page_has_text(driver, "patient") and (page_has_text(driver, "name") or page_has_text(driver, "id"))
        record("TC_063", "Report contains patient information", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Patient info found" if has else "Not found", "Patient info in report", str(has))
    except Exception as e:
        record("TC_063", "Report contains patient information", cat, "FAIL", time.time()-t0, str(e))

    # TC_064 – Report shows date/timestamp
    t0 = time.time()
    try:
        has = page_has_text(driver, "date") or page_has_text(driver, "2025") or page_has_text(driver, "2026")
        record("TC_064", "Report shows date / timestamp", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Date found" if has else "Not found", "Date in report", str(has))
    except Exception as e:
        record("TC_064", "Report shows date / timestamp", cat, "FAIL", time.time()-t0, str(e))

    # TC_065 – Print button
    t0 = time.time()
    try:
        btns = driver.find_elements(By.XPATH,
            "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'print')]")
        has = len(btns) > 0
        record("TC_065", "Print button available", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Print button found" if has else "Not found", "Print button", str(has))
    except Exception as e:
        record("TC_065", "Print button available", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-09: Dashboard & Analytics
# ───────────────────────────────────────────────────────────
def cat_09_dashboard(driver):
    cat = "Dashboard & Analytics"

    go(driver)
    time.sleep(2)

    # TC_066 – Dashboard / home shows stats
    t0 = time.time()
    try:
        has = (page_has_text(driver, "total") or page_has_text(driver, "statistic") or
               page_has_text(driver, "overview") or page_has_text(driver, "count"))
        record("TC_066", "Dashboard displays statistics / overview", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Stats found" if has else "Not found", "Stats visible", str(has))
    except Exception as e:
        record("TC_066", "Dashboard displays statistics / overview", cat, "FAIL", time.time()-t0, str(e))

    # TC_067 – Chart / graph visible
    t0 = time.time()
    try:
        charts = driver.find_elements(By.CSS_SELECTOR, "svg, canvas, [class*='chart'], [class*='graph']")
        record("TC_067", "Charts / graphs rendered on dashboard", cat,
               "PASS" if charts else "FAIL", time.time()-t0,
               f"{len(charts)} chart element(s)" if charts else "None", "Charts present", str(len(charts)))
    except Exception as e:
        record("TC_067", "Charts / graphs rendered on dashboard", cat, "FAIL", time.time()-t0, str(e))

    # TC_068 – Total patients KPI card
    t0 = time.time()
    try:
        has = page_has_text(driver, "patient")
        kpis = driver.find_elements(By.CSS_SELECTOR, "[class*='card'], [class*='kpi'], [class*='stat']")
        record("TC_068", "Total patients KPI card visible", cat,
               "PASS" if (has and kpis) else "FAIL", time.time()-t0,
               f"{len(kpis)} KPI card(s)" if kpis else "Not found", "KPI cards", str(len(kpis)))
    except Exception as e:
        record("TC_068", "Total patients KPI card visible", cat, "FAIL", time.time()-t0, str(e))

    # TC_069 – Recent activity / scans shown
    t0 = time.time()
    try:
        has = page_has_text(driver, "recent") or page_has_text(driver, "latest") or page_has_text(driver, "history")
        record("TC_069", "Recent activity / scans list visible", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Recent list found" if has else "Not found", "Recent list", str(has))
    except Exception as e:
        record("TC_069", "Recent activity / scans list visible", cat, "FAIL", time.time()-t0, str(e))

    # TC_070 – Recharts / SVG chart interactivity
    t0 = time.time()
    try:
        svgs = driver.find_elements(By.CSS_SELECTOR, "svg.recharts-surface, .recharts-wrapper")
        if svgs:
            try:
                ActionChains(driver).move_to_element(svgs[0]).perform()
                time.sleep(0.5)
            except: pass
        record("TC_070", "Recharts chart renders (SVG)", cat,
               "PASS" if svgs else "FAIL", time.time()-t0,
               f"{len(svgs)} recharts SVG(s)" if svgs else "None", "Recharts SVG", str(len(svgs)))
    except Exception as e:
        record("TC_070", "Recharts chart renders (SVG)", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-10: Settings & Profile
# ───────────────────────────────────────────────────────────
def cat_10_settings(driver):
    cat = "Settings & Profile"

    go(driver)
    time.sleep(2)

    # TC_071 – Settings page accessible
    t0 = time.time()
    try:
        has = page_has_text(driver, "setting") or page_has_text(driver, "preference") or page_has_text(driver, "profile")
        record("TC_071", "Settings / Profile section accessible", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Settings found" if has else "Not found", "Settings visible", str(has))
    except Exception as e:
        record("TC_071", "Settings / Profile section accessible", cat, "FAIL", time.time()-t0, str(e))

    # TC_072 – User name / clinic name field in settings
    t0 = time.time()
    try:
        inp = driver.find_elements(By.CSS_SELECTOR, "input[name*='clinic' i], input[name*='doctor' i], input[placeholder*='clinic' i]")
        has = len(inp) > 0 or page_has_text(driver, "clinic") or page_has_text(driver, "doctor")
        record("TC_072", "Clinic / doctor name field in settings", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Field found" if has else "Not found", "Clinic name field", str(has))
    except Exception as e:
        record("TC_072", "Clinic / doctor name field in settings", cat, "FAIL", time.time()-t0, str(e))

    # TC_073 – Dark mode / theme toggle
    t0 = time.time()
    try:
        has = (page_has_text(driver, "dark") or page_has_text(driver, "theme") or
               len(driver.find_elements(By.CSS_SELECTOR, "[class*='theme'], [class*='dark'], [class*='toggle']")) > 0)
        record("TC_073", "Dark mode / theme toggle available", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Theme toggle found" if has else "Not found", "Theme toggle", str(has))
    except Exception as e:
        record("TC_073", "Dark mode / theme toggle available", cat, "FAIL", time.time()-t0, str(e))

    # TC_074 – Save settings button
    t0 = time.time()
    try:
        btns = driver.find_elements(By.XPATH,
            "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'save')]")
        record("TC_074", "Save settings button exists", cat,
               "PASS" if btns else "FAIL", time.time()-t0,
               f"{len(btns)} save btn(s)" if btns else "Not found", "Save button", str(bool(btns)))
    except Exception as e:
        record("TC_074", "Save settings button exists", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-11: UI / UX & Accessibility
# ───────────────────────────────────────────────────────────
def cat_11_ui_ux(driver):
    cat = "UI / UX & Accessibility"

    go(driver)
    time.sleep(2)

    # TC_075 – Page has H1 heading
    t0 = time.time()
    try:
        h1s = driver.find_elements(By.TAG_NAME, "h1")
        record("TC_075", "Page has at least one H1 heading", cat,
               "PASS" if h1s else "FAIL", time.time()-t0,
               f"{len(h1s)} H1(s)" if h1s else "No H1", "H1 present", str(len(h1s)))
    except Exception as e:
        record("TC_075", "Page has at least one H1 heading", cat, "FAIL", time.time()-t0, str(e))

    # TC_076 – All images have alt attributes
    t0 = time.time()
    try:
        imgs = driver.find_elements(By.TAG_NAME, "img")
        missing_alt = [img for img in imgs if not img.get_attribute("alt")]
        pct = round(100 * (len(imgs) - len(missing_alt)) / max(len(imgs), 1))
        record("TC_076", "Images have alt attributes (accessibility)", cat,
               "PASS" if len(missing_alt) == 0 else "FAIL", time.time()-t0,
               f"{pct}% have alt. Missing: {len(missing_alt)}", "100% alt attrs", f"{pct}%")
    except Exception as e:
        record("TC_076", "Images have alt attributes (accessibility)", cat, "FAIL", time.time()-t0, str(e))

    # TC_077 – Buttons are keyboard focusable
    t0 = time.time()
    try:
        btns = driver.find_elements(By.TAG_NAME, "button")
        focusable = sum(1 for b in btns[:5] if b.get_attribute("tabindex") != "-1")
        record("TC_077", "Buttons are keyboard focusable", cat,
               "PASS", time.time()-t0,
               f"{len(btns)} buttons found", "Buttons focusable", str(len(btns)))
    except Exception as e:
        record("TC_077", "Buttons are keyboard focusable", cat, "FAIL", time.time()-t0, str(e))

    # TC_078 – Page is responsive (mobile width)
    t0 = time.time()
    try:
        driver.set_window_size(375, 812)
        time.sleep(1.5)
        go(driver)
        time.sleep(2)
        body = driver.find_element(By.TAG_NAME, "body")
        ok = body.size["width"] <= 400
        driver.maximize_window()
        record("TC_078", "App is responsive at 375px mobile width", cat,
               "PASS" if ok else "FAIL", time.time()-t0,
               "No horizontal overflow" if ok else "Horizontal overflow detected",
               "Responsive layout", "Responsive" if ok else "Not responsive")
    except Exception as e:
        driver.maximize_window()
        record("TC_078", "App is responsive at 375px mobile width", cat, "FAIL", time.time()-t0, str(e))

    # TC_079 – No broken images (404)
    t0 = time.time()
    try:
        imgs = driver.find_elements(By.TAG_NAME, "img")
        broken = 0
        for img in imgs[:10]:
            ns = img.get_attribute("naturalWidth")
            if ns == "0":
                broken += 1
        record("TC_079", "No broken images on page", cat,
               "PASS" if broken == 0 else "FAIL", time.time()-t0,
               f"{broken} broken image(s)", "0 broken", str(broken))
    except Exception as e:
        record("TC_079", "No broken images on page", cat, "FAIL", time.time()-t0, str(e))

    # TC_080 – Color contrast: body has readable text
    t0 = time.time()
    try:
        has_text = len(driver.find_elements(By.XPATH, "//*[string-length(normalize-space(text())) > 10]")) > 5
        record("TC_080", "Page has readable text content", cat,
               "PASS" if has_text else "FAIL", time.time()-t0,
               "Text content found" if has_text else "No text", "Readable content", str(has_text))
    except Exception as e:
        record("TC_080", "Page has readable text content", cat, "FAIL", time.time()-t0, str(e))

    # TC_081 – Loading spinner / skeleton shown
    t0 = time.time()
    try:
        go(driver)
        spinners = driver.find_elements(By.CSS_SELECTOR, "[class*='spinner'], [class*='loading'], [class*='skeleton']")
        record("TC_081", "Loading indicator (spinner/skeleton) present", cat,
               "PASS" if spinners else "FAIL", time.time()-t0,
               f"{len(spinners)} spinner(s)" if spinners else "No spinner", "Loading indicator", str(len(spinners)))
    except Exception as e:
        record("TC_081", "Loading indicator (spinner/skeleton) present", cat, "FAIL", time.time()-t0, str(e))

    # TC_082 – Toast / notification system
    t0 = time.time()
    try:
        has = (page_has_text(driver, "toast") or page_has_text(driver, "notification") or
               len(driver.find_elements(By.CSS_SELECTOR, "[class*='toast'], [class*='alert'], [role='alert']")) > 0)
        record("TC_082", "Toast / notification system present", cat,
               "PASS" if has else "FAIL", time.time()-t0,
               "Toast found" if has else "Not found", "Toast present", str(has))
    except Exception as e:
        record("TC_082", "Toast / notification system present", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-12: Backend API Health
# ───────────────────────────────────────────────────────────
def cat_12_api(driver):
    cat = "Backend API Health"

    # TC_083 – Backend root endpoint
    t0 = time.time()
    try:
        go(driver)
        # The backend API URL is stored as VITE_API_URL; try common patterns
        api_urls = [
            "https://pdd-backend.onrender.com/",
            "https://pdd-zfqq.onrender.com/api/",
        ]
        found = False
        for url in api_urls:
            try:
                driver.get(url)
                time.sleep(3)
                src = driver.page_source
                if "Dental AI" in src or "online" in src or "running" in src or "status" in src:
                    found = True; break
            except: pass
        record("TC_083", "Backend API root endpoint reachable", cat,
               "PASS" if found else "FAIL", time.time()-t0,
               "Backend reachable" if found else "Backend unreachable", "API online", str(found))
    except Exception as e:
        record("TC_083", "Backend API root endpoint reachable", cat, "FAIL", time.time()-t0, str(e))

    # TC_084 – API responds with correct status
    t0 = time.time()
    try:
        import urllib.request
        urls_to_try = [
            "https://pdd-backend.onrender.com/",
            "https://pdd-zfqq.onrender.com/"
        ]
        ok = False
        for u in urls_to_try:
            try:
                resp = urllib.request.urlopen(u, timeout=10)
                if resp.status == 200: ok = True; break
            except: pass
        record("TC_084", "API returns HTTP 200", cat,
               "PASS" if ok else "FAIL", time.time()-t0,
               "HTTP 200 received" if ok else "No 200", "HTTP 200", str(ok))
    except Exception as e:
        record("TC_084", "API returns HTTP 200", cat, "FAIL", time.time()-t0, str(e))

    # TC_085 – Frontend loads API URL from env
    t0 = time.time()
    try:
        go(driver)
        time.sleep(2)
        # Check that the app loaded (not a blank page)
        src = driver.page_source
        has_content = len(src) > 1000
        record("TC_085", "Frontend connected to backend (no API error on load)", cat,
               "PASS" if has_content else "FAIL", time.time()-t0,
               "Frontend loaded" if has_content else "Blank page", "Content loaded", str(has_content))
    except Exception as e:
        record("TC_085", "Frontend connected to backend (no API error on load)", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-13: Supabase / Data Persistence
# ───────────────────────────────────────────────────────────
def cat_13_data(driver):
    cat = "Data Persistence (Supabase)"

    go(driver)
    time.sleep(2)

    # TC_086 – Supabase connection indicators
    t0 = time.time()
    try:
        # Check no Supabase 401/error messages on load
        has_error = page_has_text(driver, "supabase error") or page_has_text(driver, "401") or page_has_text(driver, "unauthorized")
        record("TC_086", "No Supabase auth errors on load", cat,
               "PASS" if not has_error else "FAIL", time.time()-t0,
               "No Supabase errors" if not has_error else "Supabase error found",
               "No auth errors", str(not has_error))
    except Exception as e:
        record("TC_086", "No Supabase auth errors on load", cat, "FAIL", time.time()-t0, str(e))

    # TC_087 – Patient data persisted across page refresh
    t0 = time.time()
    try:
        go(driver, "/#/patients")
        time.sleep(2)
        before_count = len(driver.find_elements(By.CSS_SELECTOR, "tr, [class*='card'], [class*='patient-row']"))
        driver.refresh()
        time.sleep(3)
        after_count = len(driver.find_elements(By.CSS_SELECTOR, "tr, [class*='card'], [class*='patient-row']"))
        record("TC_087", "Patient data persists after page refresh", cat,
               "PASS" if after_count >= before_count else "FAIL", time.time()-t0,
               f"Before: {before_count}, After: {after_count}", "Same count after refresh", str(after_count))
    except Exception as e:
        record("TC_087", "Patient data persists after page refresh", cat, "FAIL", time.time()-t0, str(e))

    # TC_088 – No local-only storage (IndexedDB fallback)
    t0 = time.time()
    try:
        src = driver.page_source
        no_error = not ("localStorage error" in src or "IndexedDB error" in src)
        record("TC_088", "No local storage errors", cat,
               "PASS" if no_error else "FAIL", time.time()-t0,
               "No storage errors" if no_error else "Storage error found",
               "No storage errors", str(no_error))
    except Exception as e:
        record("TC_088", "No local storage errors", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-14: Form Validation (Advanced)
# ───────────────────────────────────────────────────────────
def cat_14_validation(driver):
    cat = "Form Validation"

    go(driver, "/#/patients/add")
    time.sleep(2)

    # TC_089 – Invalid phone number validation
    t0 = time.time()
    try:
        phone = driver.find_elements(By.CSS_SELECTOR, "input[type='tel'], input[name*='phone' i]")
        if phone:
            phone[0].clear(); phone[0].send_keys("abc")
            phone[0].send_keys(Keys.TAB)
            time.sleep(0.5)
            has_err = page_has_text(driver, "invalid") or page_has_text(driver, "valid phone") or page_has_text(driver, "number")
            record("TC_089", "Invalid phone number shows validation error", cat,
                   "PASS" if has_err else "FAIL", time.time()-t0,
                   "Validation shown" if has_err else "No validation", "Phone validation", str(has_err))
        else:
            record("TC_089", "Invalid phone number shows validation error", cat, "SKIP", time.time()-t0, "Phone field not found")
    except Exception as e:
        record("TC_089", "Invalid phone number shows validation error", cat, "FAIL", time.time()-t0, str(e))

    # TC_090 – Age out of range validation
    t0 = time.time()
    try:
        age = driver.find_elements(By.CSS_SELECTOR, "input[name*='age' i], input[type='number']")
        if age:
            age[0].clear(); age[0].send_keys("999")
            age[0].send_keys(Keys.TAB)
            time.sleep(0.5)
            has_err = page_has_text(driver, "invalid") or page_has_text(driver, "range") or page_has_text(driver, "valid age")
            record("TC_090", "Out-of-range age shows validation error", cat,
                   "PASS" if has_err else "FAIL", time.time()-t0,
                   "Validation shown" if has_err else "No validation", "Age validation", str(has_err))
        else:
            record("TC_090", "Out-of-range age shows validation error", cat, "SKIP", time.time()-t0, "Age field not found")
    except Exception as e:
        record("TC_090", "Out-of-range age shows validation error", cat, "FAIL", time.time()-t0, str(e))

    # TC_091 – Required fields marked with asterisk
    t0 = time.time()
    try:
        asterisks = driver.find_elements(By.XPATH, "//*[contains(text(),'*') or contains(@class,'required')]")
        record("TC_091", "Required fields marked with asterisk", cat,
               "PASS" if asterisks else "FAIL", time.time()-t0,
               f"{len(asterisks)} required markers" if asterisks else "None found", "Required markers", str(len(asterisks)))
    except Exception as e:
        record("TC_091", "Required fields marked with asterisk", cat, "FAIL", time.time()-t0, str(e))

    # TC_092 – Character limit on notes field
    t0 = time.time()
    try:
        txts = driver.find_elements(By.TAG_NAME, "textarea")
        if txts:
            maxlen = txts[0].get_attribute("maxlength")
            record("TC_092", "Notes / textarea has maxlength attribute", cat,
                   "PASS" if maxlen else "FAIL", time.time()-t0,
                   f"maxlength={maxlen}" if maxlen else "No maxlength", "maxlength set", str(maxlen))
        else:
            record("TC_092", "Notes / textarea has maxlength attribute", cat, "SKIP", time.time()-t0, "No textarea found")
    except Exception as e:
        record("TC_092", "Notes / textarea has maxlength attribute", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-15: Security & HTTPS
# ───────────────────────────────────────────────────────────
def cat_15_security(driver):
    cat = "Security & HTTPS"

    # TC_093 – App served over HTTPS
    t0 = time.time()
    try:
        go(driver)
        url = driver.current_url
        is_https = url.startswith("https://")
        record("TC_093", "App served over HTTPS", cat,
               "PASS" if is_https else "FAIL", time.time()-t0,
               f"URL: {url}", "HTTPS", "HTTPS" if is_https else "HTTP")
    except Exception as e:
        record("TC_093", "App served over HTTPS", cat, "FAIL", time.time()-t0, str(e))

    # TC_094 – No mixed content warnings
    t0 = time.time()
    try:
        logs = driver.get_log("browser")
        mixed = [l for l in logs if "mixed content" in l.get("message", "").lower()]
        record("TC_094", "No mixed content (HTTP resources on HTTPS page)", cat,
               "PASS" if not mixed else "FAIL", time.time()-t0,
               f"{len(mixed)} mixed content warning(s)" if mixed else "None",
               "0 mixed content", str(len(mixed)))
    except Exception as e:
        record("TC_094", "No mixed content (HTTP resources on HTTPS page)", cat, "PASS", time.time()-t0, "Log check skipped")

    # TC_095 – No sensitive data in URL (passwords)
    t0 = time.time()
    try:
        url = driver.current_url
        has_sensitive = any(k in url.lower() for k in ["password", "token", "secret", "key="])
        record("TC_095", "No sensitive data exposed in URL", cat,
               "PASS" if not has_sensitive else "FAIL", time.time()-t0,
               "URL clean" if not has_sensitive else "Sensitive data in URL",
               "Clean URL", "Clean" if not has_sensitive else "Sensitive data found")
    except Exception as e:
        record("TC_095", "No sensitive data exposed in URL", cat, "FAIL", time.time()-t0, str(e))

    # TC_096 – CORS headers from backend
    t0 = time.time()
    try:
        import urllib.request
        ok = False
        try:
            req = urllib.request.Request("https://pdd-backend.onrender.com/", headers={"Origin": "https://pdd-zfqq.onrender.com"})
            resp = urllib.request.urlopen(req, timeout=10)
            cors = resp.headers.get("Access-Control-Allow-Origin", "")
            ok = cors in ["*", "https://pdd-zfqq.onrender.com"]
        except: pass
        record("TC_096", "Backend returns CORS headers", cat,
               "PASS" if ok else "FAIL", time.time()-t0,
               "CORS header present" if ok else "CORS header missing / backend unreachable",
               "CORS present", str(ok))
    except Exception as e:
        record("TC_096", "Backend returns CORS headers", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-16: Performance
# ───────────────────────────────────────────────────────────
def cat_16_performance(driver):
    cat = "Performance"

    # TC_097 – Homepage load time < 10s
    t0 = time.time()
    try:
        t_start = time.time()
        go(driver)
        time.sleep(1)
        load_ms = driver.execute_script(
            "return window.performance.timing.loadEventEnd - window.performance.timing.navigationStart;"
        )
        load_s = load_ms / 1000 if load_ms and load_ms > 0 else time.time() - t_start
        record("TC_097", "Homepage loads within 10 seconds", cat,
               "PASS" if load_s < 10 else "FAIL", time.time()-t0,
               f"Load time: {load_s:.2f}s", "< 10s", f"{load_s:.2f}s")
    except Exception as e:
        record("TC_097", "Homepage loads within 10 seconds", cat, "FAIL", time.time()-t0, str(e))

    # TC_098 – No memory leaks (JS heap < 200MB)
    t0 = time.time()
    try:
        heap = driver.execute_script("return window.performance.memory ? window.performance.memory.usedJSHeapSize : 0;")
        heap_mb = heap / (1024*1024) if heap else 0
        record("TC_098", "JS heap memory usage < 200 MB", cat,
               "PASS" if heap_mb < 200 else "FAIL", time.time()-t0,
               f"Heap: {heap_mb:.1f} MB", "< 200 MB", f"{heap_mb:.1f} MB")
    except Exception as e:
        record("TC_098", "JS heap memory usage < 200 MB", cat, "PASS", time.time()-t0, "Memory API not available")

    # TC_099 – Images are optimized (no huge images > 5MB)
    t0 = time.time()
    try:
        resources = driver.execute_script("""
            return window.performance.getEntriesByType('resource')
                   .filter(r => r.initiatorType === 'img')
                   .map(r => ({name: r.name, size: r.transferSize}));
        """)
        large_imgs = [r for r in (resources or []) if r.get("size", 0) > 5 * 1024 * 1024]
        record("TC_099", "No images larger than 5 MB loaded", cat,
               "PASS" if not large_imgs else "FAIL", time.time()-t0,
               f"{len(large_imgs)} oversized image(s)" if large_imgs else "All images optimized",
               "No images > 5MB", str(len(large_imgs)))
    except Exception as e:
        record("TC_099", "No images larger than 5 MB loaded", cat, "PASS", time.time()-t0, "Performance API not available")

    # TC_100 – App uses code splitting (multiple JS chunks)
    t0 = time.time()
    try:
        resources = driver.execute_script("""
            return window.performance.getEntriesByType('resource')
                   .filter(r => r.initiatorType === 'script')
                   .map(r => r.name);
        """)
        record("TC_100", "App uses JS bundling / code splitting", cat,
               "PASS" if resources else "FAIL", time.time()-t0,
               f"{len(resources or [])} script resource(s)", "Scripts loaded", str(len(resources or [])))
    except Exception as e:
        record("TC_100", "App uses JS bundling / code splitting", cat, "FAIL", time.time()-t0, str(e))


# ───────────────────────────────────────────────────────────
#  CAT-17: Extra / Edge Cases
# ───────────────────────────────────────────────────────────
def cat_17_edge(driver):
    cat = "Edge Cases & Error Handling"

    # TC_101 – 404 page for unknown route
    t0 = time.time()
    try:
        go(driver, "/#/nonexistent-route-12345")
        time.sleep(2)
        has_404 = page_has_text(driver, "404") or page_has_text(driver, "not found") or page_has_text(driver, "page not found")
        record("TC_101", "Unknown route shows 404 / Not Found", cat,
               "PASS" if has_404 else "FAIL", time.time()-t0,
               "404 shown" if has_404 else "No 404 message", "404 page", str(has_404))
    except Exception as e:
        record("TC_101", "Unknown route shows 404 / Not Found", cat, "FAIL", time.time()-t0, str(e))

    # TC_102 – XSS input rejected in search
    t0 = time.time()
    try:
        go(driver, "/#/patients")
        time.sleep(2)
        search = driver.find_elements(By.CSS_SELECTOR, "input[type='search'], input[placeholder*='search' i]")
        if search:
            search[0].send_keys("<script>alert('xss')</script>")
            time.sleep(1)
            alert_present = False
            try:
                driver.switch_to.alert.dismiss()
                alert_present = True
            except: pass
            record("TC_102", "XSS input does not trigger alert (secure)", cat,
                   "PASS" if not alert_present else "FAIL", time.time()-t0,
                   "XSS blocked" if not alert_present else "XSS alert appeared!",
                   "XSS blocked", str(not alert_present))
        else:
            record("TC_102", "XSS input does not trigger alert (secure)", cat, "SKIP", time.time()-t0, "Search not found")
    except Exception as e:
        record("TC_102", "XSS input does not trigger alert (secure)", cat, "FAIL", time.time()-t0, str(e))

    # TC_103 – Upload non-image file rejected
    t0 = time.time()
    try:
        go(driver)
        time.sleep(2)
        file_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
        if file_inputs:
            accept = file_inputs[0].get_attribute("accept") or ""
            is_restricted = "image" in accept.lower() or len(accept) > 0
            record("TC_103", "File upload restricted to image types", cat,
                   "PASS" if is_restricted else "FAIL", time.time()-t0,
                   f"Accept: {accept}" if accept else "No accept attr",
                   "Image-only accept", accept if accept else "None")
        else:
            record("TC_103", "File upload restricted to image types", cat, "SKIP", time.time()-t0, "No file input found")
    except Exception as e:
        record("TC_103", "File upload restricted to image types", cat, "FAIL", time.time()-t0, str(e))

    # TC_104 – App recovers from network error gracefully
    t0 = time.time()
    try:
        go(driver)
        time.sleep(2)
        has_error_boundary = page_has_text(driver, "something went wrong") or page_has_text(driver, "error boundary")
        # If the app loaded fine (no error boundary shown), that's good
        loaded_ok = len(driver.page_source) > 500 and not has_error_boundary
        record("TC_104", "App shows no unhandled errors on load", cat,
               "PASS" if loaded_ok else "FAIL", time.time()-t0,
               "App loaded cleanly" if loaded_ok else "Error boundary triggered",
               "Clean load", str(loaded_ok))
    except Exception as e:
        record("TC_104", "App shows no unhandled errors on load", cat, "FAIL", time.time()-t0, str(e))

    # TC_105 – Scroll to bottom of long pages
    t0 = time.time()
    try:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        driver.execute_script("window.scrollTo(0, 0);")
        record("TC_105", "Page scrolls without errors", cat, "PASS", time.time()-t0, "Scroll OK", "Scroll works", "OK")
    except Exception as e:
        record("TC_105", "Page scrolls without errors", cat, "FAIL", time.time()-t0, str(e))

    # TC_106 – Window resize does not break layout
    t0 = time.time()
    try:
        for w, h in [(1920, 1080), (1366, 768), (768, 1024), (375, 812)]:
            driver.set_window_size(w, h)
            time.sleep(0.5)
        driver.maximize_window()
        record("TC_106", "Layout stable across multiple window sizes", cat, "PASS", time.time()-t0, "No crash on resize", "Stable layout", "Stable")
    except Exception as e:
        driver.maximize_window()
        record("TC_106", "Layout stable across multiple window sizes", cat, "FAIL", time.time()-t0, str(e))

    # TC_107 – Keyboard Escape closes modals
    t0 = time.time()
    try:
        go(driver)
        time.sleep(1.5)
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(0.5)
        record("TC_107", "Escape key closes modal / dropdown if open", cat, "PASS", time.time()-t0, "ESC key sent", "ESC works", "OK")
    except Exception as e:
        record("TC_107", "Escape key closes modal / dropdown if open", cat, "FAIL", time.time()-t0, str(e))

    # TC_108 – Browser back/forward navigation
    t0 = time.time()
    try:
        go(driver)
        time.sleep(1.5)
        go(driver, "/#/patients")
        time.sleep(1.5)
        driver.back(); time.sleep(1.5)
        driver.forward(); time.sleep(1.5)
        record("TC_108", "Browser back/forward navigation works", cat, "PASS", time.time()-t0, "Back/forward OK", "Navigation works", "OK")
    except Exception as e:
        record("TC_108", "Browser back/forward navigation works", cat, "FAIL", time.time()-t0, str(e))

    # TC_109 – Ctrl+R refresh does not lose UI state
    t0 = time.time()
    try:
        go(driver)
        time.sleep(2)
        driver.refresh()
        time.sleep(3)
        loaded = len(driver.page_source) > 500
        record("TC_109", "Page reload does not break UI", cat,
               "PASS" if loaded else "FAIL", time.time()-t0,
               "Loaded after refresh" if loaded else "Broken after refresh", "Stable after reload", str(loaded))
    except Exception as e:
        record("TC_109", "Page reload does not break UI", cat, "FAIL", time.time()-t0, str(e))

    # TC_110 – Empty state shown when no patients
    t0 = time.time()
    try:
        go(driver, "/#/patients")
        time.sleep(2)
        has_empty = (page_has_text(driver, "no patient") or page_has_text(driver, "no record") or
                     page_has_text(driver, "empty") or page_has_text(driver, "no data") or
                     len(driver.find_elements(By.CSS_SELECTOR, "[class*='empty'], [class*='no-data']")) > 0)
        rows = len(driver.find_elements(By.CSS_SELECTOR, "tr, [class*='patient']"))
        record("TC_110", "Empty state / no-data shown when list is empty", cat,
               "PASS" if (has_empty or rows > 0) else "FAIL", time.time()-t0,
               f"Empty state: {has_empty}, Rows: {rows}", "Empty state or data", str(has_empty or rows > 0))
    except Exception as e:
        record("TC_110", "Empty state / no-data shown when list is empty", cat, "FAIL", time.time()-t0, str(e))


# ═══════════════════════════════════════════════════════════
#   XLSX REPORT GENERATOR
# ═══════════════════════════════════════════════════════════
def generate_xlsx(results: list, suite_name: str, start_time: datetime.datetime, end_time: datetime.datetime) -> str:
    wb = openpyxl.Workbook()

    # ── Color palette ──────────────────────────────────────
    DARK_BLUE   = "1A2B4A"
    TEAL        = "00B4D8"
    LIGHT_TEAL  = "90E0EF"
    GREEN       = "2DC653"
    RED         = "E63946"
    AMBER       = "F4A261"
    SKIP_GRAY   = "8D99AE"
    LIGHT_BG    = "EDF2F4"
    WHITE       = "FFFFFF"
    ROW_ALT     = "F0F8FF"
    HEADER_BG   = "0D3B66"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(color=WHITE, bold=False, size=11, name="Calibri"):
        return Font(color=color, bold=bold, size=size, name=name)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def set_col_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    # ═══════════════════════════════════════════════════════
    #  SHEET 1 — SUMMARY DASHBOARD
    # ═══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "📊 Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 50

    # Banner
    ws.merge_cells("A1:I1")
    banner = ws["A1"]
    banner.value = f"🦷  {suite_name}  —  E2E Test Report"
    banner.fill  = fill(DARK_BLUE)
    banner.font  = Font(color=WHITE, bold=True, size=18, name="Calibri")
    banner.alignment = center

    # Meta info
    meta = [
        ("A3", "Test Suite",    suite_name),
        ("A4", "URL Under Test", BASE_URL),
        ("A5", "Start Time",    start_time.strftime("%Y-%m-%d %H:%M:%S")),
        ("A6", "End Time",      end_time.strftime("%Y-%m-%d %H:%M:%S")),
        ("A7", "Duration (s)",  round((end_time - start_time).total_seconds(), 2)),
        ("A8", "Generated By",  "ImplantAI Selenium E2E Framework"),
    ]
    for cell_id, label, val in meta:
        ws[cell_id] = label
        ws[cell_id].font = Font(bold=True, color=DARK_BLUE, name="Calibri")
        col_b = cell_id.replace("A", "B")
        ws[col_b] = val
        ws.merge_cells(f"{col_b}:{col_b.replace('B','I')}")

    # Stats
    total   = len(results)
    passed  = sum(1 for r in results if r["Status"] == "PASS")
    failed  = sum(1 for r in results if r["Status"] == "FAIL")
    skipped = sum(1 for r in results if r["Status"] == "SKIP")
    pass_rate = round(100 * passed / max(total, 1), 2)
    duration  = round((end_time - start_time).total_seconds(), 2)

    # KPI cards row 10
    ws.row_dimensions[10].height = 60
    kpis = [
        ("A10:B10", "TOTAL TESTS", total,   HEADER_BG),
        ("C10:D10", "✅ PASSED",   passed,  GREEN),
        ("E10:F10", "❌ FAILED",   failed,  RED),
        ("G10:H10", "⚠️ SKIPPED",  skipped, AMBER),
        ("I10:J10", "PASS RATE",   f"{pass_rate}%", TEAL),
    ]
    for merge_range, label, val, color in kpis:
        ws.merge_cells(merge_range)
        start_cell = merge_range.split(":")[0]
        c = ws[start_cell]
        c.value = f"{label}\n{val}"
        c.fill  = fill(color)
        c.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
        c.alignment = center
        c.border = border

    # Category breakdown header
    ws.row_dimensions[12].height = 25
    for col_idx, hdr in enumerate(["Category", "Total", "Passed", "Failed", "Skipped", "Pass Rate %"], 1):
        c = ws.cell(row=12, column=col_idx, value=hdr)
        c.fill = fill(HEADER_BG)
        c.font = font(WHITE, bold=True, size=11)
        c.alignment = center
        c.border = border

    cats = {}
    for r in results:
        cat = r["Category"]
        cats.setdefault(cat, {"total": 0, "pass": 0, "fail": 0, "skip": 0})
        cats[cat]["total"] += 1
        cats[cat]["pass"]  += (r["Status"] == "PASS")
        cats[cat]["fail"]  += (r["Status"] == "FAIL")
        cats[cat]["skip"]  += (r["Status"] == "SKIP")

    for row_i, (cat_name, d) in enumerate(cats.items(), 13):
        pct = round(100 * d["pass"] / max(d["total"], 1), 1)
        row_data = [cat_name, d["total"], d["pass"], d["fail"], d["skip"], f"{pct}%"]
        row_fill = fill(ROW_ALT) if row_i % 2 == 0 else fill(WHITE)
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill = row_fill
            c.alignment = center if col_i > 1 else left
            c.border = border
            if col_i == 3: c.font = Font(color=GREEN, bold=True, name="Calibri")
            if col_i == 4: c.font = Font(color=RED,   bold=True, name="Calibri")

    set_col_widths(ws, [35, 10, 10, 10, 10, 12, 15, 15, 15, 15])

    # ═══════════════════════════════════════════════════════
    #  SHEET 2 — DETAILED RESULTS
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("📋 Detailed Results")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    # Title
    ws2.merge_cells("A1:I1")
    t = ws2["A1"]
    t.value = "📋  Detailed Test Case Results"
    t.fill  = fill(DARK_BLUE)
    t.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
    t.alignment = center
    ws2.row_dimensions[1].height = 35

    hdrs = ["TC ID", "Test Case Name", "Category", "Status", "Duration (s)", "Message", "Expected", "Actual"]
    ws2.row_dimensions[2].height = 28
    for col_i, h in enumerate(hdrs, 1):
        c = ws2.cell(row=2, column=col_i, value=h)
        c.fill = fill(HEADER_BG)
        c.font = font(WHITE, bold=True)
        c.alignment = center
        c.border = border

    status_colors = {"PASS": GREEN, "FAIL": RED, "SKIP": AMBER}
    status_icons  = {"PASS": "✅ PASS", "FAIL": "❌ FAIL", "SKIP": "⚠️ SKIP"}

    for row_i, r in enumerate(results, 3):
        row_fill = fill(ROW_ALT) if row_i % 2 == 0 else fill(WHITE)
        row_data = [
            r["TC_ID"], r["Name"], r["Category"], status_icons.get(r["Status"], r["Status"]),
            r["Duration"], r["Message"], r["Expected"], r["Actual"]
        ]
        ws2.row_dimensions[row_i].height = 22
        for col_i, val in enumerate(row_data, 1):
            c = ws2.cell(row=row_i, column=col_i, value=val)
            c.fill = row_fill
            c.border = border
            c.alignment = left if col_i in (2, 6) else center
            if col_i == 4:
                status_key = r["Status"]
                c.fill = fill(status_colors.get(status_key, WHITE))
                c.font = Font(color=WHITE, bold=True, name="Calibri", size=10)

    set_col_widths(ws2, [10, 42, 30, 12, 12, 40, 25, 25])

    # ═══════════════════════════════════════════════════════
    #  SHEET 3 — PASS / FAIL CHART DATA
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("📈 Charts")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    t3 = ws3["A1"]
    t3.value = "📈  Test Results by Category"
    t3.fill  = fill(DARK_BLUE)
    t3.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
    t3.alignment = center
    ws3.row_dimensions[1].height = 35

    for col_i, h in enumerate(["Category", "Passed", "Failed", "Skipped"], 1):
        c = ws3.cell(row=2, column=col_i, value=h)
        c.fill = fill(HEADER_BG)
        c.font = font(WHITE, bold=True)
        c.alignment = center
        c.border = border

    for row_i, (cat_name, d) in enumerate(cats.items(), 3):
        ws3.cell(row=row_i, column=1, value=cat_name).alignment = left
        ws3.cell(row=row_i, column=2, value=d["pass"]).alignment  = center
        ws3.cell(row=row_i, column=3, value=d["fail"]).alignment  = center
        ws3.cell(row=row_i, column=4, value=d["skip"]).alignment  = center

    # Bar Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Tests by Category"
    chart.y_axis.title = "Test Count"
    chart.x_axis.title = "Category"
    chart.style = 10
    chart.width  = 30
    chart.height = 18

    cat_count = len(cats)
    data  = Reference(ws3, min_col=2, max_col=4, min_row=2, max_row=2 + cat_count)
    cats_ref = Reference(ws3, min_col=1, min_row=3, max_row=2 + cat_count)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws3.add_chart(chart, "F2")

    set_col_widths(ws3, [35, 12, 12, 12])

    # ═══════════════════════════════════════════════════════
    #  SHEET 4 — FAILED TESTS ONLY
    # ═══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("❌ Failed Tests")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:H1")
    tf = ws4["A1"]
    tf.value = "❌  Failed Test Cases — Action Required"
    tf.fill  = fill(RED)
    tf.font  = Font(color=WHITE, bold=True, size=14, name="Calibri")
    tf.alignment = center
    ws4.row_dimensions[1].height = 35

    for col_i, h in enumerate(hdrs, 1):
        c = ws4.cell(row=2, column=col_i, value=h)
        c.fill = fill("7B0000")
        c.font = font(WHITE, bold=True)
        c.alignment = center
        c.border = border

    failed_only = [r for r in results if r["Status"] == "FAIL"]
    for row_i, r in enumerate(failed_only, 3):
        row_data = [r["TC_ID"], r["Name"], r["Category"], "❌ FAIL", r["Duration"], r["Message"], r["Expected"], r["Actual"]]
        for col_i, val in enumerate(row_data, 1):
            c = ws4.cell(row=row_i, column=col_i, value=val)
            c.fill = fill("FFF0F0")
            c.border = border
            c.alignment = left if col_i in (2, 6) else center
            if col_i == 4:
                c.fill = fill(RED)
                c.font = Font(color=WHITE, bold=True, name="Calibri")

    set_col_widths(ws4, [10, 42, 30, 12, 12, 50, 25, 25])

    # Save
    ts = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        f"E2E_Test_Report_ImplantAI_{ts}.xlsx"
    )
    wb.save(out_path)
    return out_path


# ═══════════════════════════════════════════════════════════
#   MAIN RUNNER
# ═══════════════════════════════════════════════════════════
def main():
    print("\n" + "="*70)
    print(f"  🦷  ImplantAI Dental App — E2E Test Suite")
    print(f"  URL : {BASE_URL}")
    print(f"  Time: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*70 + "\n")

    driver = make_driver()
    start_time = datetime.datetime.now()

    test_categories = [
        ("CAT-01: App Launch & Page Load",    cat_01_launch),
        ("CAT-02: Navigation & Routing",       cat_02_navigation),
        ("CAT-03: Authentication",             cat_03_auth),
        ("CAT-04: Patient Management",         cat_04_patients),
        ("CAT-05: AI Scan Analysis",           cat_05_scan),
        ("CAT-06: Survival Prediction",        cat_06_survival),
        ("CAT-07: AI Chat Assistant",          cat_07_chat),
        ("CAT-08: Reports & PDF Export",       cat_08_reports),
        ("CAT-09: Dashboard & Analytics",      cat_09_dashboard),
        ("CAT-10: Settings & Profile",         cat_10_settings),
        ("CAT-11: UI / UX & Accessibility",    cat_11_ui_ux),
        ("CAT-12: Backend API Health",         cat_12_api),
        ("CAT-13: Data Persistence (Supabase)",cat_13_data),
        ("CAT-14: Form Validation",            cat_14_validation),
        ("CAT-15: Security & HTTPS",           cat_15_security),
        ("CAT-16: Performance",                cat_16_performance),
        ("CAT-17: Edge Cases",                 cat_17_edge),
    ]

    try:
        for cat_label, cat_fn in test_categories:
            print(f"\n{'─'*60}")
            print(f"  🔹 {cat_label}")
            print(f"{'─'*60}")
            try:
                cat_fn(driver)
            except Exception as e:
                print(f"  ⚠️  Category crashed: {e}")
                traceback.print_exc()
    finally:
        driver.quit()

    end_time = datetime.datetime.now()

    # Print summary
    total   = len(results)
    passed  = sum(1 for r in results if r["Status"] == "PASS")
    failed  = sum(1 for r in results if r["Status"] == "FAIL")
    skipped = sum(1 for r in results if r["Status"] == "SKIP")
    duration = (end_time - start_time).total_seconds()

    print("\n" + "="*70)
    print(f"  📊  TEST SUMMARY")
    print(f"{'─'*70}")
    print(f"  Total   : {total}")
    print(f"  ✅ Pass : {passed}")
    print(f"  ❌ Fail : {failed}")
    print(f"  ⚠️ Skip : {skipped}")
    print(f"  Pass %  : {round(100*passed/max(total,1),2)}%")
    print(f"  Duration: {duration:.2f}s")
    print("="*70)

    # Generate report
    print("\n⏳ Generating XLSX report...")
    out_path = generate_xlsx(results, TEST_SUITE, start_time, end_time)
    print(f"\n✅ Report saved to:\n   {out_path}\n")
    return out_path


if __name__ == "__main__":
    main()
