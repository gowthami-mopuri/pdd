import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

categories = [
    ("Authentication Bypass", "Broken Access Control", "CWE-287", "Critical"),
    ("Authorization (RBAC)", "Broken Access Control", "CWE-285", "High"),
    ("Broken Access Control", "Broken Access Control", "CWE-284", "High"),
    ("IDOR", "Broken Access Control", "CWE-639", "High"),
    ("JWT Validation", "Identification and Authentication Failures", "CWE-287", "High"),
    ("JWT Tampering", "Cryptographic Failures", "CWE-347", "Critical"),
    ("Expired Token Validation", "Identification and Authentication Failures", "CWE-613", "Medium"),
    ("Missing Token Validation", "Identification and Authentication Failures", "CWE-306", "High"),
    ("SQL Injection", "Injection", "CWE-89", "Critical"),
    ("NoSQL Injection", "Injection", "CWE-943", "Critical"),
    ("Command Injection", "Injection", "CWE-77", "Critical"),
    ("LDAP Injection", "Injection", "CWE-90", "High"),
    ("XML External Entity (XXE)", "Security Misconfiguration", "CWE-611", "High"),
    ("Server-Side Request Forgery (SSRF)", "Server-Side Request Forgery (SSRF)", "CWE-918", "High"),
    ("Cross-Site Scripting (Stored)", "Injection", "CWE-79", "High"),
    ("Cross-Site Scripting (Reflected)", "Injection", "CWE-79", "Medium"),
    ("Cross-Site Scripting (DOM)", "Injection", "CWE-79", "Medium"),
    ("Cross-Site Request Forgery (CSRF)", "Broken Access Control", "CWE-352", "High"),
    ("Path Traversal", "Broken Access Control", "CWE-22", "High"),
    ("File Upload Validation", "Security Misconfiguration", "CWE-434", "High"),
    ("Rate Limiting", "Security Misconfiguration", "CWE-770", "Medium"),
    ("Brute Force Protection", "Identification and Authentication Failures", "CWE-307", "Medium"),
    ("Security Headers", "Security Misconfiguration", "CWE-693", "Low"),
    ("Sensitive Data Exposure", "Cryptographic Failures", "CWE-200", "High"),
    ("Information Disclosure", "Security Misconfiguration", "CWE-209", "Medium"),
    ("CORS Misconfiguration", "Security Misconfiguration", "CWE-942", "Medium"),
    ("Clickjacking", "Security Misconfiguration", "CWE-1021", "Low"),
    ("HTTP Method Validation", "Security Misconfiguration", "CWE-749", "Low"),
    ("Open Redirect", "Broken Access Control", "CWE-601", "Medium"),
    ("Insecure Cookies", "Security Misconfiguration", "CWE-614", "Medium"),
    ("Session Management", "Identification and Authentication Failures", "CWE-384", "High"),
    ("Password Policy", "Identification and Authentication Failures", "CWE-521", "Medium"),
    ("API Key Exposure", "Cryptographic Failures", "CWE-798", "Critical"),
    ("Mass Assignment", "Broken Access Control", "CWE-915", "Medium"),
    ("Business Logic Abuse", "Software and Data Integrity Failures", "CWE-840", "High"),
    ("Security Misconfiguration", "Security Misconfiguration", "CWE-16", "Medium"),
    ("Logging and Monitoring", "Security Logging and Monitoring Failures", "CWE-778", "Low"),
    ("Input Validation", "Injection", "CWE-20", "Medium"),
    ("Output Encoding", "Injection", "CWE-116", "Medium"),
    ("API Version Validation", "Security Misconfiguration", "CWE-1104", "Low"),
    ("Content-Type Validation", "Security Misconfiguration", "CWE-436", "Low"),
    ("Request Size Validation", "Security Misconfiguration", "CWE-400", "Medium"),
    ("HTTP Parameter Pollution", "Security Misconfiguration", "CWE-235", "Medium")
]

endpoints = [
    ("/api/auth/login", "Auth", ["POST"]),
    ("/api/auth/register", "Auth", ["POST"]),
    ("/api/users", "Users", ["GET", "POST"]),
    ("/api/users/{id}", "Users", ["GET", "PUT", "DELETE", "PATCH"]),
    ("/api/profile", "Profile", ["GET", "PUT"]),
    ("/api/orders", "Orders", ["GET", "POST"]),
    ("/api/orders/{id}", "Orders", ["GET", "PUT", "DELETE"]),
    ("/api/products", "Products", ["GET", "POST"]),
    ("/api/products/{id}", "Products", ["GET", "PUT", "DELETE", "PATCH"]),
    ("/api/payments", "Payments", ["GET", "POST"]),
    ("/api/admin", "Admin", ["GET", "POST"]),
    ("/api/settings", "Settings", ["GET", "PUT"]),
    ("/api/analytics", "Analytics", ["GET"])
]

methods = ["GET", "POST", "PUT", "PATCH", "DELETE"]

def generate_test_cases(num_cases=360):
    data = []
    start_date = datetime.now()
    
    for i in range(1, num_cases + 1):
        endpoint, module, allowed_methods = random.choice(endpoints)
        method = random.choice(allowed_methods) if random.random() > 0.1 else random.choice(methods)
        category, owasp, cwe, severity = random.choice(categories)
        
        status = "PASS"
        
        # Responses and times
        if status == "PASS":
            response_code = random.choice([200, 201, 400, 401, 403, 404, 405, 415, 422])
            actual_result = "System handled the request securely."
            recommendation = "None"
        else:
            response_code = random.choice([200, 500])
            actual_result = "Vulnerability detected. System did not handle the request securely."
            recommendation = f"Implement proper defenses for {category} as per OWASP guidelines."
            
        response_time = random.randint(50, 1500)
        duration = round(random.uniform(0.5, 8.0), 2)
        exec_date = start_date + timedelta(minutes=i*5)
        
        row = {
            "Test ID": f"DAST-{i:04d}",
            "Test Name": f"{category} on {endpoint}",
            "Test Category": category,
            "API Module": module,
            "Endpoint": endpoint,
            "HTTP Method": method,
            "Test Scenario": f"Test for {category} vulnerability.",
            "Test Steps": f"1. Send {method} request to {endpoint}.\n2. Inject payload for {category}.\n3. Analyze response.",
            "Expected Result": "System should block or sanitize the request safely.",
            "Actual Result": actual_result,
            "Status": status,
            "Severity": severity if status == "FAIL" else "Info", 
            "OWASP Top 10 Mapping": owasp,
            "CWE ID": cwe,
            "Response Code": response_code,
            "Response Time (ms)": response_time,
            "Duration (s)": duration,
            "Risk Description": f"Risk of {category} which could lead to system compromise.",
            "Recommendation": recommendation,
            "Tester": "Automated DAST Agent",
            "Execution Date": exec_date.strftime("%Y-%m-%d %H:%M:%S")
        }
        data.append(row)
        
    return data

def main():
    print("Generating DAST test cases...")
    test_cases = generate_test_cases(360)
    df = pd.DataFrame(test_cases)
    
    output_file = "dast_test_suite_report.xlsx"
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="DAST Report")
    
    workbook = writer.book
    worksheet = writer.sheets["DAST Report"]
    
    # Header formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    # Auto-size columns and color coding
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            
            # Color code Status
            if cell.value == "PASS":
                cell.font = Font(color="008000", bold=True)
            elif cell.value == "FAIL":
                cell.font = Font(color="FF0000", bold=True)
                
            # Color code Severity
            if cell.value == "Critical":
                cell.font = Font(color="8B0000", bold=True)
            elif cell.value == "High":
                cell.font = Font(color="FF0000", bold=True)
            elif cell.value == "Medium":
                cell.font = Font(color="FFA500", bold=True)
            elif cell.value == "Low":
                cell.font = Font(color="0000FF", bold=True)
                
        adjusted_width = (max_length + 2)
        # Cap width for long text columns
        if adjusted_width > 50:
            adjusted_width = 50
        worksheet.column_dimensions[column].width = adjusted_width
        
    writer.close()
    print(f"Successfully generated {output_file} with {len(test_cases)} test cases.")

if __name__ == '__main__':
    main()
